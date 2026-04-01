# reports/invoice_lines.py
import io
import csv
import json
import requests
from typing import Any
from calendar import monthrange
from datetime import date
from decimal import Decimal, InvalidOperation
from collections import defaultdict
from db import get_conn
from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse, JSONResponse

router = APIRouter(prefix="/reports/invoice_lines", tags=["reports-invoice-lines"])


def ref_value(ref: dict | None) -> str:
    if isinstance(ref, dict):
        return (ref.get("value") or "").strip()
    return ""

def ref_name(ref: dict | None) -> str:
    if isinstance(ref, dict):
        return (ref.get("name") or "").strip()
    return ""

def ref_json(ref: dict | None) -> str:
    # Optional: if you want a JSON string in CSV too
    import json
    return json.dumps(ref, ensure_ascii=False) if isinstance(ref, dict) else ""

# Excel Helpers
def freeze_header_row(ws):
    ws.freeze_panes = "A2"

def autosize_worksheet_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter

        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

def detail_row_for_excel(r: dict, include_customer: bool) -> dict:
    row = {
        "FamilyCode": r.get("FamilyCode", ""),
        "DocNumber": r.get("DocNumber", ""),
        "TxnDate": r.get("TxnDate", ""),
        "P.O. Number": r.get("P.O. Number", ""),
        "Amount": r.get("Amount", ""),
        "Description": r.get("Description", ""),
        "Work Order": r.get("Work Order", ""),
        "Item": r.get("Item", ""),
        "Unit Price": r.get("Unit Price", ""),
        "Qty": r.get("Qty", ""),
    }

    if include_customer:
        row["CustomerName"] = r.get("CustomerName", "")

    return row

def qualifies_robert(customer_name: str, item_code: str) -> bool:
    # Combo 1 you provided
    if customer_name.strip().lower() == "air france":
        return item_code in {"S906-70196-3", "362A6411P4"}
    # TODO: add Robert combo #2 later
    return False

def qualifies_evert(customer_name: str, item_code: str) -> bool:
    cn = (customer_name or "").lower()

    # Evert customer substring list (includes KLM; overlap is fine)
    keywords = ["ajw technique", "klm", "aar", "lufthansa", "csi", "austrian", "fokker", "ametek", "muirhead"]
    if any(k in cn for k in keywords):
        return True

    # Skysmart special cases: for now, just include all Skysmart lines,
    # OR restrict to certain items later (you can decide).
    if "skysmart" in cn:
        return True

    return False

def extract_work_order(description: str) -> str:
    """
    Extracts Work Order value from multi-line description.

    Looks for: "IRT WO#:" and returns text after it up to end-of-line.
    If not found, returns "".
    """
    if not description:
        return ""

    marker = "IRT WO#:"
    idx = description.find(marker)
    if idx == -1:
        return ""

    # Take everything after the marker
    after = description[idx + len(marker):]

    # If description is multi-line, take only the first line after the marker
    # (work order is typically on the same line)
    first_line = after.splitlines()[0] if after else ""

    return first_line.strip()

def to_decimal(val: Any) -> Decimal:
    if val in (None, "", " "):
        return Decimal("0")
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")

def fetch_item_family_map(request: Request, item_ids: list[str]) -> dict[str, str]:
    """
    Looks up family_code in item_catalog using qbo_item_id
    through the existing Postgres connection.
    """
    clean_ids = sorted({str(x).strip() for x in item_ids if str(x).strip()})
    if not clean_ids:
        return {}

    family_map: dict[str, str] = {}

    chunk_size = 500
    for i in range(0, len(clean_ids), chunk_size):
        chunk = clean_ids[i:i + chunk_size]

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    select qbo_item_id, family_code
                    from item_catalog
                    where qbo_item_id = any(%s)
                    """,
                    (chunk,)
                )
                rows = cur.fetchall() or []

                for r in rows:
                    item_id = str(r.get("qbo_item_id", "")).strip()
                    family_code = (r.get("family_code") or "").strip()
                    if item_id:
                        family_map[item_id] = family_code or "UNASSIGNED"

    return family_map

def attach_family_codes(request: Request, all_lines: list[dict]) -> list[dict]:
    item_ids = [
        str(r.get("ItemId", "")).strip()
        for r in all_lines
        if str(r.get("ItemId", "")).strip()
    ]

    family_map = fetch_item_family_map(request, item_ids)

    for r in all_lines:
        item_id = str(r.get("ItemId", "")).strip()
        r["FamilyCode"] = family_map.get(item_id, "UNASSIGNED")

    return all_lines

def group_lines_by_family(all_lines: list[dict], include_customer: bool = False) -> list[dict]:
    grouped: dict[tuple, dict] = {}

    for r in all_lines:
        customer_name = str(r.get("CustomerName", "")).strip()
        family_code = str(r.get("FamilyCode", "UNASSIGNED")).strip() or "UNASSIGNED"

        # When include_customer=True, grouping happens within each customer
        key = (customer_name, family_code) if include_customer else (family_code,)

        if key not in grouped:
            grouped[key] = {
                "CustomerName": customer_name,
                "FamilyCode": family_code,
                "Item": str(r.get("Item", "")).strip(),
                "TotalQty": Decimal("0"),
                "TotalSales": Decimal("0"),
            }

        grouped[key]["TotalQty"] += to_decimal(r.get("Qty"))
        grouped[key]["TotalSales"] += to_decimal(r.get("Amount"))

    results: list[dict] = []
    for g in grouped.values():
        if include_customer:
            results.append({
                "CustomerName": g["CustomerName"],
                "FamilyCode": g["FamilyCode"],
                "Item": g["Item"],
                "TotalQty": float(g["TotalQty"]),
                "TotalSales": float(g["TotalSales"]),
            })
        else:
            results.append({
                "FamilyCode": g["FamilyCode"],
                "Item": g["Item"],
                "TotalQty": float(g["TotalQty"]),
                "TotalSales": float(g["TotalSales"]),
            })

    if include_customer:
        results.sort(key=lambda x: (x["CustomerName"], x["FamilyCode"], x["Item"]))
    else:
        results.sort(key=lambda x: (x["FamilyCode"], x["Item"]))

    return results

def build_invoice_query(start_date: str, end_date: str, customer_id: str | None = None) -> str:
    query = (
        f"SELECT * FROM Invoice "
        f"WHERE TxnDate >= '{start_date}' "
        f"AND TxnDate <= '{end_date}'"
    )

    if customer_id is not None and str(customer_id).strip():
        safe_customer_id = str(customer_id).strip()
        query += f" AND CustomerRef = '{safe_customer_id}'"

    query += " ORDER BY TxnDate ASC, Id ASC"

    return query

def qbo_query_all(
    realm_id: str,
    query: str,
    access_token: str,
    qbo_api_base: str,
    page_size: int = 1000,
) -> list[dict]:
    """
    Runs a QBO query and fetches ALL pages using GET with `query=` param.
    This avoids QBO's common "QueryParserError: null" issue seen with POST bodies.
    """
    results: list[dict] = []
    start = 1

    while True:
        paged_query = f"{query} STARTPOSITION {start} MAXRESULTS {page_size}"
        url = f"{qbo_api_base}/v3/company/{realm_id}/query"

        r = requests.get(
            url,
            params={"query": paged_query, "minorversion": "75"},
            headers={
                "Authorization": f"Bearer {access_token}",
                "Accept": "application/json",
            },
            timeout=60,
        )

        if r.status_code == 401:
            raise RuntimeError(f"AUTH_401: {r.text}")
        if r.status_code >= 400:
            # helpful debug (no tokens leaked)
            raise RuntimeError(
                f"QBO_QUERY_FAILED ({r.status_code}): {r.text} | sent_query={paged_query}"
            )

        payload = r.json() if r.content else {}
        q = payload.get("QueryResponse", {})
        batch = q.get("Invoice", []) or []
        results.extend(batch)

        if len(batch) < page_size:
            break

        start += page_size

    return results

def safe_filter_invoices_by_customer(invoices: list, customer_id: str | None = None) -> list:
    """
    Safety filter applied after QBO returns invoices.
    Keeps all invoices if customer_id is not provided.
    """
    if customer_id is None or not str(customer_id).strip():
        return invoices

    target_customer_id = str(customer_id).strip()

    filtered = []
    for invoice in invoices:
        customer_ref = invoice.get("CustomerRef") or {}
        invoice_customer_id = str(customer_ref.get("value", "")).strip()

        if invoice_customer_id == target_customer_id:
            filtered.append(invoice)

    return filtered


def safe_json(val: Any) -> str:
    # For CSV: flatten nested dict/list safely into a JSON string
    return json.dumps(val, ensure_ascii=False, default=str) if val is not None else ""


def flatten_invoice_lines(invoice: dict) -> list[dict]:
    rows: list[dict] = []

    # Parent invoice fields requested
    invoice_id = invoice.get("Id", "")
    doc_number = invoice.get("DocNumber", "")
    txn_date = invoice.get("TxnDate", "")

    customer_ref = invoice.get("CustomerRef") or {}
    customer_name = customer_ref.get("name", "")
    customer_id = customer_ref.get("value", "")
    
    
    # Extract P.O. Number custom field value from CustomField[]
    po_number = ""
    custom_fields = invoice.get("CustomField") or []
    if isinstance(custom_fields, list):
        for cf in custom_fields:
            if not isinstance(cf, dict):
                continue
            # QBO often stores the label in Name (or DefinitionId ties to config)
            if (cf.get("Name") or "").strip().lower() in ("p.o. number", "po number", "p.o. #", "po #"):
                # Value can be in StringValue for text custom fields
                po_number = cf.get("StringValue") or cf.get("value") or ""
                break

    # Extract Terms
    sales_term_ref = invoice.get("SalesTermRef") or {}
    sales_term_name = sales_term_ref.get("name", "")    
    
    # keep the rest as you already had
    meta = invoice.get("MetaData") or {}
    lines = invoice.get("Line") or []
    if not isinstance(lines, list):
        return rows

    for idx, line in enumerate(lines, start=1):
        # ... (no change to your line parsing)

        # ONLY keep SalesItemLineDetail lines
        if line.get("DetailType") != "SalesItemLineDetail":
            continue

        # Extract SalesItemLineDetail
        sales_item_line_detail = line.get("SalesItemLineDetail") or {}
        item_ref = sales_item_line_detail.get("ItemRef") or {}
        item_name = item_ref.get("name", "") or ""
        item_id = item_ref.get("value", "")
        
        # Remove leading "FAA Repair:" if present
        prefix = "FAA Repair:"
        clean_name = item_name.strip()
        if clean_name.startswith(prefix):
            item_name = clean_name[len(prefix):].strip()
        else:
            item_name = clean_name
    
        # Force large numeric-looking Item values to be treated as text in Excel
        if item_name.isdigit() and len(item_name) >= 10:
            item_name = f'="{item_name}"'
        
        unit_price = sales_item_line_detail.get("UnitPrice", "")
        qty = sales_item_line_detail.get("Qty", "")

        line_amount = line.get("Amount", "")
        if to_decimal(line_amount) == 0:
            continue

        # Extract WO # when Description has it
        descr = line.get("Description", "") or ""
        work_order = extract_work_order(descr)

        row = {
            # Requested parent fields:
            "InvoiceId": invoice_id,
            "DocNumber": doc_number,
            "TxnDate": txn_date,
            "CustomerId": customer_id,
            "CustomerName": customer_name,
            "P.O. Number": po_number,
                        
            # Line identifiers / ordering
            "LineId": line.get("Id", ""),

            # keep the rest of your fields...
            "Amount": line_amount,
            "Description": descr,

            # Work Order number when existing in record
            "Work Order": work_order,

            "ItemId": item_id,
            "Item": item_name,
            "Unit Price": unit_price,
            "Qty": qty,
           
        }

        rows.append(row)

    return rows


@router.get("/year/{year}")
def download_invoice_lines_for_year(
    request: Request,
    realmId: str,
    year: int,
    format: str = "json",
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    # 1) Get valid token (auto-refresh inside your existing helper)
    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    # 2) Query all invoices in the year, sorted by TxnDate asc, Id asc
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)
    
    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    # 3) Flatten lines
    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)
   
    # 4) Return JSON (default) or CSV
    if format.lower() == "json":
        buf = io.BytesIO()
        buf.write(json.dumps(all_lines, indent=2, ensure_ascii=False, default=str).encode("utf-8"))
        buf.seek(0)

        if customer_id:
            filename = f"invoice_lines_{year}_{realmId}_customer_{customer_id}.json"
        else:
            filename = f"invoice_lines_{year}_{realmId}.json"
        
        return StreamingResponse(
            buf,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if format.lower() == "csv":
        customer_name = ""
        if customer_id and all_lines:
            customer_name = str(all_lines[0].get("CustomerName", "")).strip()
    
        period_label = f"Year {year}"
    
        text_buf = io.StringIO()
    
        if customer_id:
            text_buf.write(f"{customer_name}\n")
            text_buf.write(f"{period_label}\n\n")
        else:
            text_buf.write(f"{period_label}\n\n")
        
        # For “all”, CSV needs a stable set of columns.
        # We'll keep key extracted columns + JSON blobs for the complex parts.
        if customer_id:
            fieldnames = [
                # requested invoice parent fields
                #"InvoiceId",
                "DocNumber",
                "TxnDate",
                "P.O. Number",
                        
                # rest of your line fields
                "LineId",
                "Amount",
                "Description",
                "Work Order",
                "Item",
                "Unit Price",
                "Qty",
            ]
        else:
            fieldnames = [
                # requested invoice parent fields
                #"InvoiceId",
                "DocNumber",
                "TxnDate",
                "CustomerName",
                "P.O. Number",
                        
                # rest of your line fields
                "LineId",
                "Amount",
                "Description",
                "Work Order",
                "Item",
                "Unit Price",
                "Qty",
            ]

        text_buf = io.StringIO()
        writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for r in all_lines:
            row = dict(r)
            row["Line_json"] = safe_json(r.get("Line_json"))
            row["Invoice_json"] = safe_json(r.get("Invoice_json"))
            writer.writerow(row)

        data = text_buf.getvalue().encode("utf-8-sig")
        buf = io.BytesIO(data)
        buf.seek(0)

        if customer_id:
            filename = f"invoice_lines_{year}_{realmId}_customer_{customer_id}.csv"
        else:
            filename = f"invoice_lines_{year}_{realmId}.csv"

        return StreamingResponse(
            buf,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return JSONResponse({"error": "invalid_format", "allowed": ["json", "csv"]}, status_code=400)


@router.get("/month/{year}/{month}")
def download_invoice_lines_for_month(
    request: Request,
    realmId: str,
    year: int,
    month: int,
    format: str = "json",
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    # 1) Get valid token
    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    # Validate month
    if month < 1 or month > 12:
        return JSONResponse({"error": "invalid_month"}, status_code=400)

    # 2) Build correct start/end dates for that month
    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    # 3) Flatten lines
    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        all_lines.extend(rows)

    # 4) Return JSON (default) or CSV
    if format.lower() == "json":
        buf = io.BytesIO()
        buf.write(json.dumps(all_lines, indent=2, ensure_ascii=False, default=str).encode("utf-8"))
        buf.seek(0)
        
        if customer_id:
            filename = f"invoice_lines_{year}_{month:02d}_{realmId}_customer_{customer_id}.json"
        else:
            filename = f"invoice_lines_{year}_{month:02d}_{realmId}.json"
        
        return StreamingResponse(
            buf,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # CSV (same fieldnames as your year report)
    if format.lower() == "csv":
        customer_name = ""
        if customer_id and all_lines:
            customer_name = str(all_lines[0].get("CustomerName", "")).strip()
    
        period_label = f"{start_date} - {end_date}"
    
        text_buf = io.StringIO()
    
        if customer_id:
            text_buf.write(f"{customer_name}\n")
            text_buf.write(f"{period_label}\n\n")
        else:
            text_buf.write(f"{period_label}\n\n")
        
        if customer_id:
            fieldnames = [
                "DocNumber",
                "TxnDate",
                "P.O. Number",
                "LineId",
                "Amount",
                "Description",
                "Work Order",
                "Item",
                "Unit Price",
                "Qty",
            ]
        else: 
            fieldnames = [
                "DocNumber",
                "TxnDate",
                "CustomerName",
                "P.O. Number",
                "LineId",
                "Amount",
                "Description",
                "Work Order",
                "Item",
                "Unit Price",
                "Qty",
            ]

        text_buf = io.StringIO()
        writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for r in all_lines:
            writer.writerow(r)

        data = text_buf.getvalue().encode("utf-8-sig")
        buf = io.BytesIO(data)
        buf.seek(0)

        if customer_id:
            filename = f"invoice_lines_{year}_{month:02d}_{realmId}_customer_{customer_id}.csv"
        else:
            filename = f"invoice_lines_{year}_{month:02d}_{realmId}.csv"
        
        return StreamingResponse(
            buf,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return JSONResponse({"error": "invalid_format"}, status_code=400)

@router.get("/quarter/{year}/{quarter}")
def download_invoice_lines_for_quarter(
    request: Request,
    realmId: str,
    year: int,
    quarter: int,
    format: str = "json",
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if quarter not in (1, 2, 3, 4):
        return JSONResponse({"error": "invalid_quarter", "allowed": [1, 2, 3, 4]}, status_code=400)

    if quarter == 1:
        start_date = f"{year}-01-01"
        end_date = f"{year}-03-31"
    elif quarter == 2:
        start_date = f"{year}-04-01"
        end_date = f"{year}-06-30"
    elif quarter == 3:
        start_date = f"{year}-07-01"
        end_date = f"{year}-09-30"
    else:
        start_date = f"{year}-10-01"
        end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    if format.lower() == "json":
        buf = io.BytesIO()
        buf.write(json.dumps(all_lines, indent=2, ensure_ascii=False, default=str).encode("utf-8"))
        buf.seek(0)

        if customer_id:
            filename = f"invoice_lines_{year}_Q{quarter}_{realmId}_customer_{customer_id}.json"
        else:
            filename = f"invoice_lines_{year}_Q{quarter}_{realmId}.json"

        return StreamingResponse(
            buf,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if format.lower() == "csv":
        customer_name = ""
        if customer_id and all_lines:
            customer_name = str(all_lines[0].get("CustomerName", "")).strip()
    
        period_label = f"{year} Q{quarter}"
    
        text_buf = io.StringIO()
    
        if customer_id:
            text_buf.write(f"{customer_name}\n")
            text_buf.write(f"{period_label}\n\n")
        else:
            text_buf.write(f"{period_label}\n\n")
        
        if customer_id:
            fieldnames = [
                "DocNumber",
                "TxnDate",
                "P.O. Number",
                "LineId",
                "Amount",
                "Description",
                "Work Order",
                "Item",
                "Unit Price",
                "Qty",
            ]
        else:
            fieldnames = [
                "DocNumber",
                "TxnDate",
                "CustomerName",
                "P.O. Number",
                "LineId",
                "Amount",
                "Description",
                "Work Order",
                "Item",
                "Unit Price",
                "Qty",
            ]            

        text_buf = io.StringIO()
        writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for r in all_lines:
            writer.writerow(r)

        data = text_buf.getvalue().encode("utf-8-sig")
        buf = io.BytesIO(data)
        buf.seek(0)

        if customer_id:
            filename = f"invoice_lines_{year}_Q{quarter}_{realmId}_customer_{customer_id}.csv"
        else:
            filename = f"invoice_lines_{year}_Q{quarter}_{realmId}.csv"

        return StreamingResponse(
            buf,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return JSONResponse({"error": "invalid_format", "allowed": ["json", "csv"]}, status_code=400)

@router.get("/grouped/year/{year}")
def download_invoice_lines_grouped_by_family_for_year(
    request: Request,
    realmId: str,
    year: int,
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)
    grouped_rows = group_lines_by_family(all_lines, include_customer=(customer_id is None))
    
    customer_name = ""
    if customer_id and all_lines:
        customer_name = str(all_lines[0].get("CustomerName", "")).strip()

    text_buf = io.StringIO()

    if customer_id:
        text_buf.write(f"{customer_name}\n")
        text_buf.write(f"Year {year}\n\n")
    else:
        text_buf.write(f"Year {year}\n\n")

    if customer_id:
        fieldnames = [
            "FamilyCode",
            "Item",
            "TotalQty",
            "TotalSales",
        ]
    else:
        fieldnames = [
            "CustomerName",
            "FamilyCode",
            "Item",
            "TotalQty",
            "TotalSales",
        ]

    writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for r in grouped_rows:
        writer.writerow(r)

    data = text_buf.getvalue().encode("utf-8-sig")
    buf = io.BytesIO(data)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_grouped_family_{year}_{realmId}_customer_{customer_id}.csv"
    else:
        filename = f"invoice_lines_grouped_family_{year}_{realmId}.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/grouped/month/{year}/{month}")
def download_invoice_lines_grouped_by_family_for_month(
    request: Request,
    realmId: str,
    year: int,
    month: int,
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if month < 1 or month > 12:
        return JSONResponse({"error": "invalid_month"}, status_code=400)

    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)
    grouped_rows = group_lines_by_family(all_lines, include_customer=(customer_id is None))
    
    customer_name = ""
    if customer_id and all_lines:
        customer_name = str(all_lines[0].get("CustomerName", "")).strip()

    period_label = f"{year}-{month:02d}"

    text_buf = io.StringIO()

    if customer_id:
        text_buf.write(f"{customer_name}\n")
        text_buf.write(f"{period_label}\n\n")
    else:
        text_buf.write(f"{period_label}\n\n")

    if customer_id:
        fieldnames = [
            "FamilyCode",
            "Item",
            "TotalQty",
            "TotalSales",
        ]
    else:
        fieldnames = [
            "CustomerName",
            "FamilyCode",
            "Item",
            "TotalQty",
            "TotalSales",
        ]

    writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for r in grouped_rows:
        writer.writerow(r)

    data = text_buf.getvalue().encode("utf-8-sig")
    buf = io.BytesIO(data)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_grouped_family_{year}_{month:02d}_{realmId}_customer_{customer_id}.csv"
    else:
        filename = f"invoice_lines_grouped_family_{year}_{month:02d}_{realmId}.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/grouped/quarter/{year}/{quarter}")
def download_invoice_lines_grouped_by_family_for_quarter(
    request: Request,
    realmId: str,
    year: int,
    quarter: int,
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if quarter not in (1, 2, 3, 4):
        return JSONResponse({"error": "invalid_quarter", "allowed": [1, 2, 3, 4]}, status_code=400)

    if quarter == 1:
        start_date = f"{year}-01-01"
        end_date = f"{year}-03-31"
    elif quarter == 2:
        start_date = f"{year}-04-01"
        end_date = f"{year}-06-30"
    elif quarter == 3:
        start_date = f"{year}-07-01"
        end_date = f"{year}-09-30"
    else:
        start_date = f"{year}-10-01"
        end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)
    grouped_rows = group_lines_by_family(all_lines, include_customer=(customer_id is None))

    customer_name = ""
    if customer_id and all_lines:
        customer_name = str(all_lines[0].get("CustomerName", "")).strip()

    period_label = f"{year} Q{quarter}"

    text_buf = io.StringIO()

    if customer_id:
        text_buf.write(f"{customer_name}\n")
        text_buf.write(f"{period_label}\n\n")
    else:
        text_buf.write(f"{period_label}\n\n")

    if customer_id:
        fieldnames = [
            "FamilyCode",
            "Item",
            "TotalQty",
            "TotalSales",
        ]
    else:
        fieldnames = [
            "CustomerName",
            "FamilyCode",
            "Item",
            "TotalQty",
            "TotalSales",
        ]

    writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for r in grouped_rows:
        writer.writerow(r)

    data = text_buf.getvalue().encode("utf-8-sig")
    buf = io.BytesIO(data)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_grouped_family_{year}_Q{quarter}_{realmId}_customer_{customer_id}.csv"
    else:
        filename = f"invoice_lines_grouped_family_{year}_Q{quarter}_{realmId}.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/grouped-lines/year/{year}")
def download_invoice_lines_with_family_for_year(
    request: Request,
    realmId: str,
    year: int,
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)  # already excludes Amount = 0
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    # 🔥 attach family (NO grouping)
    all_lines = attach_family_codes(request, all_lines)
    
    all_lines.sort(key=lambda x: (
    str(x.get("CustomerName", "")) if customer_id is None else "",
    str(x.get("FamilyCode", "")),
    # str(x.get("Item", "")),
    str(x.get("TxnDate", "")),
    ))

    # CSV output
    customer_name = ""
    if customer_id and all_lines:
        customer_name = str(all_lines[0].get("CustomerName", "")).strip()

    text_buf = io.StringIO()

    if customer_id:
        text_buf.write(f"{customer_name}\n")
        text_buf.write(f"Year {year}\n\n")
    else:
        text_buf.write(f"Year {year}\n\n")
    
    if customer_id:
        fieldnames = [
            "FamilyCode",   # 👈 NEW FIELD
            "DocNumber",
            "TxnDate",
            "P.O. Number",
            "Amount",
            "Description",
            "Work Order",
            "Item",
            "Unit Price",
            "Qty",
        ]
    else:
        fieldnames = [
            "FamilyCode",   # 👈 NEW FIELD
            "DocNumber",
            "TxnDate",
            "CustomerName",
            "P.O. Number",
            "Amount",
            "Description",
            "Work Order",
            "Item",
            "Unit Price",
            "Qty",
        ]

    writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for r in all_lines:
        writer.writerow(r)

    data = text_buf.getvalue().encode("utf-8-sig")
    buf = io.BytesIO(data)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_with_family_{year}_{realmId}_customer_{customer_id}.csv"
    else:
        filename = f"invoice_lines_with_family_{year}_{realmId}.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/grouped-lines/month/{year}/{month}")
def download_invoice_lines_with_family_for_month(
    request: Request,
    realmId: str,
    year: int,
    month: int,
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if month < 1 or month > 12:
        return JSONResponse({"error": "invalid_month"}, status_code=400)

    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)

    all_lines.sort(key=lambda x: (
        str(x.get("CustomerName", "")) if customer_id is None else "",
        str(x.get("FamilyCode", "")),
        # str(x.get("Item", "")),
        str(x.get("TxnDate", "")),
    ))

    # CSV output
    customer_name = ""
    if customer_id and all_lines:
        customer_name = str(all_lines[0].get("CustomerName", "")).strip()

    text_buf = io.StringIO()

    if customer_id:
        text_buf.write(f"{customer_name}\n")
        text_buf.write(f"{start_date} - {end_date}\n\n")
    else:
        text_buf.write(f"{start_date} - {end_date}\n\n")
    
    if customer_id:
        fieldnames = [
            "FamilyCode",
            "DocNumber",
            "TxnDate",
            "P.O. Number",
            "Amount",
            "Description",
            "Work Order",
            "Item",
            "Unit Price",
            "Qty",
        ]
    else:
        fieldnames = [
            "FamilyCode",
            "DocNumber",
            "TxnDate",
            "CustomerName",
            "P.O. Number",
            "Amount",
            "Description",
            "Work Order",
            "Item",
            "Unit Price",
            "Qty",
        ]        

    # text_buf = io.StringIO()
    writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for r in all_lines:
        writer.writerow(r)

    data = text_buf.getvalue().encode("utf-8-sig")
    buf = io.BytesIO(data)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_with_family_{year}_{month:02d}_{realmId}_customer_{customer_id}.csv"
    else:
        filename = f"invoice_lines_with_family_{year}_{month:02d}_{realmId}.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/grouped-lines/quarter/{year}/{quarter}")
def download_invoice_lines_with_family_for_quarter(
    request: Request,
    realmId: str,
    year: int,
    quarter: int,
    customer_id: str | None = None,
):
    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if quarter not in (1, 2, 3, 4):
        return JSONResponse({"error": "invalid_quarter"}, status_code=400)

    if quarter == 1:
        start_date = f"{year}-01-01"
        end_date = f"{year}-03-31"
    elif quarter == 2:
        start_date = f"{year}-04-01"
        end_date = f"{year}-06-30"
    elif quarter == 3:
        start_date = f"{year}-07-01"
        end_date = f"{year}-09-30"
    else:
        start_date = f"{year}-10-01"
        end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)

    all_lines.sort(key=lambda x: (
        str(x.get("CustomerName", "")) if customer_id is None else "",
        str(x.get("FamilyCode", "")),
        # str(x.get("Item", "")),
        str(x.get("TxnDate", "")),
    ))

    # CSV output
    customer_name = ""
    if customer_id and all_lines:
        customer_name = str(all_lines[0].get("CustomerName", "")).strip()

    text_buf = io.StringIO()

    if customer_id:
        text_buf.write(f"{customer_name}\n")
        text_buf.write(f"{start_date} - {end_date}\n\n")
    else:
        text_buf.write(f"{start_date} - {end_date}\n\n")
    
    if customer_id:
        fieldnames = [
            "FamilyCode",
            "DocNumber",
            "TxnDate",
            "P.O. Number",
            "Amount",
            "Description",
            "Work Order",
            "Item",
            "Unit Price",
            "Qty",
        ]
    else:
        fieldnames = [
            "FamilyCode",
            "DocNumber",
            "TxnDate",
            "CustomerName",
            "P.O. Number",
            "Amount",
            "Description",
            "Work Order",
            "Item",
            "Unit Price",
            "Qty",
        ]        

    # text_buf = io.StringIO()
    writer = csv.DictWriter(text_buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for r in all_lines:
        writer.writerow(r)

    data = text_buf.getvalue().encode("utf-8-sig")
    buf = io.BytesIO(data)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_with_family_{year}_Q{quarter}_{realmId}_customer_{customer_id}.csv"
    else:
        filename = f"invoice_lines_with_family_{year}_Q{quarter}_{realmId}.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/excel/year/{year}")
def download_invoice_lines_excel_for_year(
    request: Request,
    realmId: str,
    year: int,
    customer_id: str | None = None,
):
    from openpyxl import Workbook
    from collections import defaultdict

    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)

    # Sorting
    all_lines.sort(key=lambda x: (
        str(x.get("CustomerName", "")) if customer_id is None else "",
        str(x.get("FamilyCode", "")),
        str(x.get("TxnDate", "")),
    ))

    grouped_rows = group_lines_by_family(all_lines, include_customer=(customer_id is None))

    # Create workbook
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # --- Summary sheet ---
    if grouped_rows:
        headers = list(grouped_rows[0].keys())
        ws_summary.append(headers)
        for row in grouped_rows:
            ws_summary.append([row.get(h, "") for h in headers])
    
    freeze_header_row(ws_summary)
    autosize_worksheet_columns(ws_summary)

    # --- Detail sheets ---
    if customer_id:
        ws_detail = wb.create_sheet(title="Detail")
    
        detail_rows = [detail_row_for_excel(r, include_customer=False) for r in all_lines]
    
        if detail_rows:
            headers = list(detail_rows[0].keys())
            ws_detail.append(headers)
            for row in detail_rows:
                ws_detail.append([row.get(h, "") for h in headers])
    
        freeze_header_row(ws_detail)
        autosize_worksheet_columns(ws_detail)
    else:
        by_customer = defaultdict(list)
    
        for r in all_lines:
            name = (r.get("CustomerName") or "UNKNOWN")[:31]
            by_customer[name].append(r)
    
        for customer_name, rows in by_customer.items():
            ws = wb.create_sheet(title=customer_name)
    
            detail_rows = [detail_row_for_excel(r, include_customer=False) for r in rows]
    
            if detail_rows:
                headers = list(detail_rows[0].keys())
                ws.append(headers)
                for row in detail_rows:
                    ws.append([row.get(h, "") for h in headers])
    
            freeze_header_row(ws)
            autosize_worksheet_columns(ws)

    # Save file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_excel_{year}_{realmId}_customer_{customer_id}.xlsx"
    else:
        filename = f"invoice_lines_excel_{year}_{realmId}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/excel/month/{year}/{month}")
def download_invoice_lines_excel_for_month(
    request: Request,
    realmId: str,
    year: int,
    month: int,
    customer_id: str | None = None,
):
    from openpyxl import Workbook
    from collections import defaultdict

    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if month < 1 or month > 12:
        return JSONResponse({"error": "invalid_month"}, status_code=400)

    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)

    all_lines.sort(key=lambda x: (
        str(x.get("CustomerName", "")) if customer_id is None else "",
        str(x.get("FamilyCode", "")),
        str(x.get("TxnDate", "")),
    ))

    grouped_rows = group_lines_by_family(all_lines, include_customer=(customer_id is None))

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    if grouped_rows:
        headers = list(grouped_rows[0].keys())
        ws_summary.append(headers)
        for row in grouped_rows:
            ws_summary.append([row.get(h, "") for h in headers])
    
    freeze_header_row(ws_summary)
    autosize_worksheet_columns(ws_summary)    

    if customer_id:
        ws_detail = wb.create_sheet(title="Detail")
    
        detail_rows = [detail_row_for_excel(r, include_customer=False) for r in all_lines]
    
        if detail_rows:
            headers = list(detail_rows[0].keys())
            ws_detail.append(headers)
            for row in detail_rows:
                ws_detail.append([row.get(h, "") for h in headers])
    
        freeze_header_row(ws_detail)
        autosize_worksheet_columns(ws_detail)
    else:
        by_customer = defaultdict(list)
    
        for r in all_lines:
            name = (r.get("CustomerName") or "UNKNOWN")[:31]
            by_customer[name].append(r)
    
        for customer_name, rows in by_customer.items():
            ws = wb.create_sheet(title=customer_name)
    
            detail_rows = [detail_row_for_excel(r, include_customer=False) for r in rows]
    
            if detail_rows:
                headers = list(detail_rows[0].keys())
                ws.append(headers)
                for row in detail_rows:
                    ws.append([row.get(h, "") for h in headers])
    
            freeze_header_row(ws)
            autosize_worksheet_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"invoice_lines_excel_{year}_{month:02d}_{realmId}_customer_{customer_id}.xlsx"
        if customer_id
        else f"invoice_lines_excel_{year}_{month:02d}_{realmId}.xlsx"
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/excel/quarter/{year}/{quarter}")
def download_invoice_lines_excel_for_quarter(
    request: Request,
    realmId: str,
    year: int,
    quarter: int,
    customer_id: str | None = None,
):
    from openpyxl import Workbook
    from collections import defaultdict

    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    if quarter not in (1, 2, 3, 4):
        return JSONResponse({"error": "invalid_quarter"}, status_code=400)

    if quarter == 1:
        start_date = f"{year}-01-01"
        end_date = f"{year}-03-31"
    elif quarter == 2:
        start_date = f"{year}-04-01"
        end_date = f"{year}-06-30"
    elif quarter == 3:
        start_date = f"{year}-07-01"
        end_date = f"{year}-09-30"
    else:
        start_date = f"{year}-10-01"
        end_date = f"{year}-12-31"

    q = build_invoice_query(start_date, end_date, customer_id=customer_id)

    invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
    invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

    all_lines: list[dict] = []
    for inv in invoices:
        rows = flatten_invoice_lines(inv)
        for r in rows:
            r["RealmId"] = realmId
        all_lines.extend(rows)

    all_lines = attach_family_codes(request, all_lines)

    all_lines.sort(key=lambda x: (
        str(x.get("CustomerName", "")) if customer_id is None else "",
        str(x.get("FamilyCode", "")),
        str(x.get("TxnDate", "")),
    ))

    grouped_rows = group_lines_by_family(all_lines, include_customer=(customer_id is None))

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    if grouped_rows:
        headers = list(grouped_rows[0].keys())
        ws_summary.append(headers)
        for row in grouped_rows:
            ws_summary.append([row.get(h, "") for h in headers])
    
    freeze_header_row(ws_summary)
    autosize_worksheet_columns(ws_summary)

    if customer_id:
        ws_detail = wb.create_sheet(title="Detail")
    
        detail_rows = [detail_row_for_excel(r, include_customer=False) for r in all_lines]
    
        if detail_rows:
            headers = list(detail_rows[0].keys())
            ws_detail.append(headers)
            for row in detail_rows:
                ws_detail.append([row.get(h, "") for h in headers])
    
        freeze_header_row(ws_detail)
        autosize_worksheet_columns(ws_detail)
    else:
        by_customer = defaultdict(list)
    
        for r in all_lines:
            name = (r.get("CustomerName") or "UNKNOWN")[:31]
            by_customer[name].append(r)
    
        for customer_name, rows in by_customer.items():
            ws = wb.create_sheet(title=customer_name)
    
            detail_rows = [detail_row_for_excel(r, include_customer=False) for r in rows]
    
            if detail_rows:
                headers = list(detail_rows[0].keys())
                ws.append(headers)
                for row in detail_rows:
                    ws.append([row.get(h, "") for h in headers])
    
            freeze_header_row(ws)
            autosize_worksheet_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"invoice_lines_excel_{year}_Q{quarter}_{realmId}_customer_{customer_id}.xlsx"
        if customer_id
        else f"invoice_lines_excel_{year}_Q{quarter}_{realmId}.xlsx"
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/compare/year/{year_a}/vs/{year_b}")
def compare_invoice_lines_summary_year_vs_year(
    request: Request,
    realmId: str,
    year_a: int,
    year_b: int,
    customer_id: str | None = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    def fetch_period_lines(year_val: int) -> list[dict]:
        start_date = f"{year_val}-01-01"
        end_date = f"{year_val}-12-31"

        q = build_invoice_query(start_date, end_date, customer_id=customer_id)
        invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
        invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

        lines: list[dict] = []
        for inv in invoices:
            rows = flatten_invoice_lines(inv)  # already excludes Amount = 0
            for r in rows:
                r["RealmId"] = realmId
            lines.extend(rows)

        lines = attach_family_codes(request, lines)
        return lines

    def summarize_by_family(lines: list[dict]) -> dict[str, dict]:
        summary: dict[str, dict] = {}

        for r in lines:
            family_code = str(r.get("FamilyCode", "UNASSIGNED")).strip() or "UNASSIGNED"
            item_name = str(r.get("Item", "")).strip()

            if family_code not in summary:
                summary[family_code] = {
                    "FamilyCode": family_code,
                    "Item": item_name,
                    "TotalQty": Decimal("0"),
                    "TotalSales": Decimal("0"),
                }

            summary[family_code]["TotalQty"] += to_decimal(r.get("Qty"))
            summary[family_code]["TotalSales"] += to_decimal(r.get("Amount"))

        return summary

    lines_a = fetch_period_lines(year_a)
    lines_b = fetch_period_lines(year_b)

    summary_a = summarize_by_family(lines_a)
    summary_b = summarize_by_family(lines_b)

    all_families = sorted(set(summary_a.keys()) | set(summary_b.keys()))

    comparison_rows: list[dict] = []
    total_qty_a = Decimal("0")
    total_qty_b = Decimal("0")
    total_sales_a = Decimal("0")
    total_sales_b = Decimal("0")

    for family_code in all_families:
        row_a = summary_a.get(family_code, {})
        row_b = summary_b.get(family_code, {})

        item_name = (
            str(row_a.get("Item", "")).strip()
            or str(row_b.get("Item", "")).strip()
        )

        qty_a = to_decimal(row_a.get("TotalQty"))
        qty_b = to_decimal(row_b.get("TotalQty"))
        sales_a = to_decimal(row_a.get("TotalSales"))
        sales_b = to_decimal(row_b.get("TotalSales"))

        sales_diff = sales_a - sales_b
        pct_diff = None
        if sales_b != 0:
            pct_diff = (sales_diff / sales_b) * Decimal("100")

        comparison_rows.append({
            "FamilyCode": family_code,
            "Item": item_name,
            f"{year_a} UNITS": float(qty_a),
            f"{year_b} UNITS": float(qty_b),
            f"{year_a} SALES": float(sales_a),
            f"{year_b} SALES": float(sales_b),
            "$ Difference": float(sales_diff),
            "% Difference": float(pct_diff) if pct_diff is not None else None,
        })

        total_qty_a += qty_a
        total_qty_b += qty_b
        total_sales_a += sales_a
        total_sales_b += sales_b

    total_diff = total_sales_a - total_sales_b
    total_pct_diff = None
    if total_sales_b != 0:
        total_pct_diff = (total_diff / total_sales_b) * Decimal("100")

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    title = f"{year_a} v {year_b} Orders by Item Family"
    if customer_id and lines_a:
        customer_name = str(lines_a[0].get("CustomerName", "")).strip()
        if customer_name:
            title = f"{customer_name} - {title}"

    headers = [
        "FamilyCode",
        "Item",
        f"{year_a} UNITS",
        f"{year_b} UNITS",
        f"{year_a} SALES",
        f"{year_b} SALES",
        "$ Difference",
        "% Difference",
    ]

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=title)
    ws.cell(row=2, column=1).font = title_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    start_data_row = header_row + 1
    row_idx = start_data_row

    for row in comparison_rows:
        ws.cell(row=row_idx, column=1, value=row["FamilyCode"])
        ws.cell(row=row_idx, column=2, value=row["Item"])
        ws.cell(row=row_idx, column=3, value=row[f"{year_a} UNITS"])
        ws.cell(row=row_idx, column=4, value=row[f"{year_b} UNITS"])
        ws.cell(row=row_idx, column=5, value=row[f"{year_a} SALES"])
        ws.cell(row=row_idx, column=6, value=row[f"{year_b} SALES"])
        ws.cell(row=row_idx, column=7, value=row["$ Difference"])
        ws.cell(row=row_idx, column=8, value=row["% Difference"])

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border

        row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=3, value=float(total_qty_a))
    ws.cell(row=total_row, column=4, value=float(total_qty_b))
    ws.cell(row=total_row, column=5, value=float(total_sales_a))
    ws.cell(row=total_row, column=6, value=float(total_sales_b))
    ws.cell(row=total_row, column=7, value=float(total_diff))
    ws.cell(row=total_row, column=8, value=float(total_pct_diff) if total_pct_diff is not None else None)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = bold_font
        cell.border = border
        cell.fill = header_fill

    for r in range(start_data_row, total_row + 1):
        for c in [5, 6, 7]:
            ws.cell(row=r, column=c).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
        ws.cell(row=r, column=8).number_format = '0%'

    # If you want 33 instead of 0.33, use this instead of the line above:
    for r in range(start_data_row, total_row + 1):
        pct_cell = ws.cell(row=r, column=8)
        if pct_cell.value is not None:
            pct_cell.value = pct_cell.value / 100
        pct_cell.number_format = '0%'

    # Freeze header row
    ws.freeze_panes = "A5"

    # Auto-size columns
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 28)

    # Totals / commentary block
    notes_row = total_row + 3
    ws.cell(row=notes_row, column=1, value="Totals").font = bold_font
    ws.cell(row=notes_row + 1, column=1, value=f"{year_a}:")
    ws.cell(row=notes_row + 1, column=2, value=float(total_qty_a))
    ws.cell(row=notes_row + 1, column=3, value=float(total_sales_a))
    ws.cell(row=notes_row + 2, column=1, value=f"{year_b}:")
    ws.cell(row=notes_row + 2, column=2, value=float(total_qty_b))
    ws.cell(row=notes_row + 2, column=3, value=float(total_sales_b))
    ws.cell(row=notes_row + 3, column=1, value="Difference:")
    ws.cell(row=notes_row + 3, column=2, value=float(total_diff))
    if total_pct_diff is not None:
        ws.cell(row=notes_row + 3, column=3, value=float(total_pct_diff) / 100)
        ws.cell(row=notes_row + 3, column=3).number_format = '0%'

    ws.cell(row=notes_row + 1, column=3).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
    ws.cell(row=notes_row + 2, column=3).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
    ws.cell(row=notes_row + 3, column=2).number_format = '$ #,##0.00;[Red]-$ #,##0.00'

    analysis_text = (
        f"{year_a} total sales were "
        f"{'higher than' if total_diff > 0 else 'lower than' if total_diff < 0 else 'equal to'} "
        f"{year_b} by ${abs(float(total_diff)):,.2f}."
    )
    ws.cell(row=notes_row + 5, column=1, value="Analysis").font = bold_font
    ws.cell(row=notes_row + 6, column=1, value=analysis_text)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if customer_id:
        filename = f"invoice_lines_compare_{year_a}_vs_{year_b}_{realmId}_customer_{customer_id}.xlsx"
    else:
        filename = f"invoice_lines_compare_{year_a}_vs_{year_b}_{realmId}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/compare/month/{year_a}/{month_a}/vs/{year_b}/{month_b}")
def compare_invoice_lines_summary_month_vs_month(
    request: Request,
    realmId: str,
    year_a: int,
    month_a: int,
    year_b: int,
    month_b: int,
    customer_id: str | None = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    if month_a < 1 or month_a > 12 or month_b < 1 or month_b > 12:
        return JSONResponse({"error": "invalid_month"}, status_code=400)

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    def fetch_period_lines(year_val: int, month_val: int) -> list[dict]:
        last_day = monthrange(year_val, month_val)[1]
        start_date = f"{year_val}-{month_val:02d}-01"
        end_date = f"{year_val}-{month_val:02d}-{last_day:02d}"

        q = build_invoice_query(start_date, end_date, customer_id=customer_id)
        invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
        invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

        lines: list[dict] = []
        for inv in invoices:
            rows = flatten_invoice_lines(inv)
            for r in rows:
                r["RealmId"] = realmId
            lines.extend(rows)

        lines = attach_family_codes(request, lines)
        return lines

    def summarize_by_family(lines: list[dict]) -> dict[str, dict]:
        summary: dict[str, dict] = {}

        for r in lines:
            family_code = str(r.get("FamilyCode", "UNASSIGNED")).strip() or "UNASSIGNED"
            item_name = str(r.get("Item", "")).strip()

            if family_code not in summary:
                summary[family_code] = {
                    "FamilyCode": family_code,
                    "Item": item_name,
                    "TotalQty": Decimal("0"),
                    "TotalSales": Decimal("0"),
                }

            summary[family_code]["TotalQty"] += to_decimal(r.get("Qty"))
            summary[family_code]["TotalSales"] += to_decimal(r.get("Amount"))

        return summary

    lines_a = fetch_period_lines(year_a, month_a)
    lines_b = fetch_period_lines(year_b, month_b)

    summary_a = summarize_by_family(lines_a)
    summary_b = summarize_by_family(lines_b)

    all_families = sorted(set(summary_a.keys()) | set(summary_b.keys()))

    label_a = f"{year_a}-{month_a:02d}"
    label_b = f"{year_b}-{month_b:02d}"

    comparison_rows: list[dict] = []
    total_qty_a = Decimal("0")
    total_qty_b = Decimal("0")
    total_sales_a = Decimal("0")
    total_sales_b = Decimal("0")

    for family_code in all_families:
        row_a = summary_a.get(family_code, {})
        row_b = summary_b.get(family_code, {})

        item_name = (
            str(row_a.get("Item", "")).strip()
            or str(row_b.get("Item", "")).strip()
        )

        qty_a = to_decimal(row_a.get("TotalQty"))
        qty_b = to_decimal(row_b.get("TotalQty"))
        sales_a = to_decimal(row_a.get("TotalSales"))
        sales_b = to_decimal(row_b.get("TotalSales"))

        sales_diff = sales_a - sales_b
        pct_diff = None
        if sales_b != 0:
            pct_diff = (sales_diff / sales_b) * Decimal("100")

        comparison_rows.append({
            "FamilyCode": family_code,
            "Item": item_name,
            f"{label_a} UNITS": float(qty_a),
            f"{label_b} UNITS": float(qty_b),
            f"{label_a} SALES": float(sales_a),
            f"{label_b} SALES": float(sales_b),
            "$ Difference": float(sales_diff),
            "% Difference": float(pct_diff) if pct_diff is not None else None,
        })

        total_qty_a += qty_a
        total_qty_b += qty_b
        total_sales_a += sales_a
        total_sales_b += sales_b

    total_diff = total_sales_a - total_sales_b
    total_pct_diff = None
    if total_sales_b != 0:
        total_pct_diff = (total_diff / total_sales_b) * Decimal("100")

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    title = f"{label_a} v {label_b} Orders by Item Family"
    customer_name = ""
    if customer_id:
        source_lines = lines_a if lines_a else lines_b
        if source_lines:
            customer_name = str(source_lines[0].get("CustomerName", "")).strip()
        if customer_name:
            title = f"{customer_name} - {title}"

    headers = [
        "FamilyCode",
        "Item",
        f"{label_a} UNITS",
        f"{label_b} UNITS",
        f"{label_a} SALES",
        f"{label_b} SALES",
        "$ Difference",
        "% Difference",
    ]

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=title)
    ws.cell(row=2, column=1).font = title_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    start_data_row = header_row + 1
    row_idx = start_data_row

    for row in comparison_rows:
        ws.cell(row=row_idx, column=1, value=row["FamilyCode"])
        ws.cell(row=row_idx, column=2, value=row["Item"])
        ws.cell(row=row_idx, column=3, value=row[f"{label_a} UNITS"])
        ws.cell(row=row_idx, column=4, value=row[f"{label_b} UNITS"])
        ws.cell(row=row_idx, column=5, value=row[f"{label_a} SALES"])
        ws.cell(row=row_idx, column=6, value=row[f"{label_b} SALES"])
        ws.cell(row=row_idx, column=7, value=row["$ Difference"])
        ws.cell(row=row_idx, column=8, value=row["% Difference"])

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border

        row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=3, value=float(total_qty_a))
    ws.cell(row=total_row, column=4, value=float(total_qty_b))
    ws.cell(row=total_row, column=5, value=float(total_sales_a))
    ws.cell(row=total_row, column=6, value=float(total_sales_b))
    ws.cell(row=total_row, column=7, value=float(total_diff))
    ws.cell(row=total_row, column=8, value=float(total_pct_diff) if total_pct_diff is not None else None)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = bold_font
        cell.border = border
        cell.fill = header_fill

    for r in range(start_data_row, total_row + 1):
        for c in [5, 6, 7]:
            ws.cell(row=r, column=c).number_format = '$ #,##0.00;[Red]-$ #,##0.00'

        pct_cell = ws.cell(row=r, column=8)
        if pct_cell.value is not None:
            pct_cell.value = pct_cell.value / 100
        pct_cell.number_format = '0%'

    ws.freeze_panes = "A5"

    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 28)

    notes_row = total_row + 3
    ws.cell(row=notes_row, column=1, value="Totals").font = bold_font
    ws.cell(row=notes_row + 1, column=1, value=f"{label_a}:")
    ws.cell(row=notes_row + 1, column=2, value=float(total_qty_a))
    ws.cell(row=notes_row + 1, column=3, value=float(total_sales_a))
    ws.cell(row=notes_row + 2, column=1, value=f"{label_b}:")
    ws.cell(row=notes_row + 2, column=2, value=float(total_qty_b))
    ws.cell(row=notes_row + 2, column=3, value=float(total_sales_b))
    ws.cell(row=notes_row + 3, column=1, value="Difference:")
    ws.cell(row=notes_row + 3, column=2, value=float(total_diff))
    if total_pct_diff is not None:
        ws.cell(row=notes_row + 3, column=3, value=float(total_pct_diff) / 100)
        ws.cell(row=notes_row + 3, column=3).number_format = '0%'

    ws.cell(row=notes_row + 1, column=3).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
    ws.cell(row=notes_row + 2, column=3).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
    ws.cell(row=notes_row + 3, column=2).number_format = '$ #,##0.00;[Red]-$ #,##0.00'

    analysis_text = (
        f"{label_a} total sales were "
        f"{'higher than' if total_diff > 0 else 'lower than' if total_diff < 0 else 'equal to'} "
        f"{label_b} by ${abs(float(total_diff)):,.2f}."
    )
    ws.cell(row=notes_row + 5, column=1, value="Analysis").font = bold_font
    ws.cell(row=notes_row + 6, column=1, value=analysis_text)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if customer_id:
        filename = (
            f"invoice_lines_compare_{year_a}_{month_a:02d}_vs_{year_b}_{month_b:02d}_"
            f"{realmId}_customer_{customer_id}.xlsx"
        )
    else:
        filename = (
            f"invoice_lines_compare_{year_a}_{month_a:02d}_vs_{year_b}_{month_b:02d}_"
            f"{realmId}.xlsx"
        )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/compare/quarter/{year_a}/{quarter_a}/vs/{year_b}/{quarter_b}")
def compare_invoice_lines_summary_quarter_vs_quarter(
    request: Request,
    realmId: str,
    year_a: int,
    quarter_a: int,
    year_b: int,
    quarter_b: int,
    customer_id: str | None = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    get_valid_access_token = request.app.state.get_valid_access_token
    qbo_api_base = request.app.state.qbo_api_base

    if customer_id is not None:
        customer_id = customer_id.strip()
        if customer_id == "":
            return JSONResponse(
                {"error": "invalid_customer_id", "message": "customer_id cannot be empty"},
                status_code=400,
            )

    if quarter_a not in (1, 2, 3, 4) or quarter_b not in (1, 2, 3, 4):
        return JSONResponse({"error": "invalid_quarter"}, status_code=400)

    try:
        access_token = get_valid_access_token(realmId)
    except RuntimeError as e:
        msg = str(e)
        if msg.startswith("RECONNECT_REQUIRED"):
            return JSONResponse(
                {"error": "reconnect_required", "connect_url": "/connect", "message": msg},
                status_code=401,
            )
        return JSONResponse({"error": "auth_failed", "message": msg}, status_code=500)

    def fetch_period_lines(year_val: int, quarter_val: int) -> list[dict]:
        if quarter_val == 1:
            start_date = f"{year_val}-01-01"
            end_date = f"{year_val}-03-31"
        elif quarter_val == 2:
            start_date = f"{year_val}-04-01"
            end_date = f"{year_val}-06-30"
        elif quarter_val == 3:
            start_date = f"{year_val}-07-01"
            end_date = f"{year_val}-09-30"
        else:
            start_date = f"{year_val}-10-01"
            end_date = f"{year_val}-12-31"

        q = build_invoice_query(start_date, end_date, customer_id=customer_id)
        invoices = qbo_query_all(realmId, q, access_token, qbo_api_base)
        invoices = safe_filter_invoices_by_customer(invoices, customer_id=customer_id)

        lines: list[dict] = []
        for inv in invoices:
            rows = flatten_invoice_lines(inv)
            for r in rows:
                r["RealmId"] = realmId
            lines.extend(rows)

        lines = attach_family_codes(request, lines)
        return lines

    def summarize_by_family(lines: list[dict]) -> dict[str, dict]:
        summary: dict[str, dict] = {}

        for r in lines:
            family_code = str(r.get("FamilyCode", "UNASSIGNED")).strip() or "UNASSIGNED"
            item_name = str(r.get("Item", "")).strip()

            if family_code not in summary:
                summary[family_code] = {
                    "FamilyCode": family_code,
                    "Item": item_name,
                    "TotalQty": Decimal("0"),
                    "TotalSales": Decimal("0"),
                }

            summary[family_code]["TotalQty"] += to_decimal(r.get("Qty"))
            summary[family_code]["TotalSales"] += to_decimal(r.get("Amount"))

        return summary

    lines_a = fetch_period_lines(year_a, quarter_a)
    lines_b = fetch_period_lines(year_b, quarter_b)

    summary_a = summarize_by_family(lines_a)
    summary_b = summarize_by_family(lines_b)

    all_families = sorted(set(summary_a.keys()) | set(summary_b.keys()))

    label_a = f"{year_a} Q{quarter_a}"
    label_b = f"{year_b} Q{quarter_b}"

    comparison_rows: list[dict] = []
    total_qty_a = Decimal("0")
    total_qty_b = Decimal("0")
    total_sales_a = Decimal("0")
    total_sales_b = Decimal("0")

    for family_code in all_families:
        row_a = summary_a.get(family_code, {})
        row_b = summary_b.get(family_code, {})

        item_name = (
            str(row_a.get("Item", "")).strip()
            or str(row_b.get("Item", "")).strip()
        )

        qty_a = to_decimal(row_a.get("TotalQty"))
        qty_b = to_decimal(row_b.get("TotalQty"))
        sales_a = to_decimal(row_a.get("TotalSales"))
        sales_b = to_decimal(row_b.get("TotalSales"))

        sales_diff = sales_a - sales_b
        pct_diff = None
        if sales_b != 0:
            pct_diff = (sales_diff / sales_b) * Decimal("100")

        comparison_rows.append({
            "FamilyCode": family_code,
            "Item": item_name,
            f"{label_a} UNITS": float(qty_a),
            f"{label_b} UNITS": float(qty_b),
            f"{label_a} SALES": float(sales_a),
            f"{label_b} SALES": float(sales_b),
            "$ Difference": float(sales_diff),
            "% Difference": float(pct_diff) if pct_diff is not None else None,
        })

        total_qty_a += qty_a
        total_qty_b += qty_b
        total_sales_a += sales_a
        total_sales_b += sales_b

    total_diff = total_sales_a - total_sales_b
    total_pct_diff = None
    if total_sales_b != 0:
        total_pct_diff = (total_diff / total_sales_b) * Decimal("100")

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    title = f"{label_a} v {label_b} Orders by Item Family"
    customer_name = ""
    if customer_id:
        source_lines = lines_a if lines_a else lines_b
        if source_lines:
            customer_name = str(source_lines[0].get("CustomerName", "")).strip()
        if customer_name:
            title = f"{customer_name} - {title}"

    headers = [
        "FamilyCode",
        "Item",
        f"{label_a} UNITS",
        f"{label_b} UNITS",
        f"{label_a} SALES",
        f"{label_b} SALES",
        "$ Difference",
        "% Difference",
    ]

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=title)
    ws.cell(row=2, column=1).font = title_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    start_data_row = header_row + 1
    row_idx = start_data_row

    for row in comparison_rows:
        ws.cell(row=row_idx, column=1, value=row["FamilyCode"])
        ws.cell(row=row_idx, column=2, value=row["Item"])
        ws.cell(row=row_idx, column=3, value=row[f"{label_a} UNITS"])
        ws.cell(row=row_idx, column=4, value=row[f"{label_b} UNITS"])
        ws.cell(row=row_idx, column=5, value=row[f"{label_a} SALES"])
        ws.cell(row=row_idx, column=6, value=row[f"{label_b} SALES"])
        ws.cell(row=row_idx, column=7, value=row["$ Difference"])
        ws.cell(row=row_idx, column=8, value=row["% Difference"])

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border

        row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=3, value=float(total_qty_a))
    ws.cell(row=total_row, column=4, value=float(total_qty_b))
    ws.cell(row=total_row, column=5, value=float(total_sales_a))
    ws.cell(row=total_row, column=6, value=float(total_sales_b))
    ws.cell(row=total_row, column=7, value=float(total_diff))
    ws.cell(row=total_row, column=8, value=float(total_pct_diff) if total_pct_diff is not None else None)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = bold_font
        cell.border = border
        cell.fill = header_fill

    for r in range(start_data_row, total_row + 1):
        for c in [5, 6, 7]:
            ws.cell(row=r, column=c).number_format = '$ #,##0.00;[Red]-$ #,##0.00'

        pct_cell = ws.cell(row=r, column=8)
        if pct_cell.value is not None:
            pct_cell.value = pct_cell.value / 100
        pct_cell.number_format = '0%'

    ws.freeze_panes = "A5"

    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 28)

    notes_row = total_row + 3
    ws.cell(row=notes_row, column=1, value="Totals").font = bold_font
    ws.cell(row=notes_row + 1, column=1, value=f"{label_a}:")
    ws.cell(row=notes_row + 1, column=2, value=float(total_qty_a))
    ws.cell(row=notes_row + 1, column=3, value=float(total_sales_a))
    ws.cell(row=notes_row + 2, column=1, value=f"{label_b}:")
    ws.cell(row=notes_row + 2, column=2, value=float(total_qty_b))
    ws.cell(row=notes_row + 2, column=3, value=float(total_sales_b))
    ws.cell(row=notes_row + 3, column=1, value="Difference:")
    ws.cell(row=notes_row + 3, column=2, value=float(total_diff))
    if total_pct_diff is not None:
        ws.cell(row=notes_row + 3, column=3, value=float(total_pct_diff) / 100)
        ws.cell(row=notes_row + 3, column=3).number_format = '0%'

    ws.cell(row=notes_row + 1, column=3).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
    ws.cell(row=notes_row + 2, column=3).number_format = '$ #,##0.00;[Red]-$ #,##0.00'
    ws.cell(row=notes_row + 3, column=2).number_format = '$ #,##0.00;[Red]-$ #,##0.00'

    analysis_text = (
        f"{label_a} total sales were "
        f"{'higher than' if total_diff > 0 else 'lower than' if total_diff < 0 else 'equal to'} "
        f"{label_b} by ${abs(float(total_diff)):,.2f}."
    )
    ws.cell(row=notes_row + 5, column=1, value="Analysis").font = bold_font
    ws.cell(row=notes_row + 6, column=1, value=analysis_text)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if customer_id:
        filename = (
            f"invoice_lines_compare_{year_a}_Q{quarter_a}_vs_{year_b}_Q{quarter_b}_"
            f"{realmId}_customer_{customer_id}.xlsx"
        )
    else:
        filename = (
            f"invoice_lines_compare_{year_a}_Q{quarter_a}_vs_{year_b}_Q{quarter_b}_"
            f"{realmId}.xlsx"
        )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )
